"""
Stock Guide FII
gerar_dados.py | Extrai dados do Excel e gera data/fiis.js

Uso:
  py -3 scripts/gerar_dados.py

Dependencias:
  pip install openpyxl
"""

import glob
import json
import os
import sys

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'Data')
OUTPUT = os.path.join(os.path.dirname(__file__), '..', 'data', 'fiis.js')

xlsx_files = sorted(glob.glob(os.path.join(DATA_DIR, '*.xlsx')))
if not xlsx_files:
    print('ERRO: Nenhum arquivo .xlsx encontrado em /Data')
    sys.exit(1)

XLSX_PATH = xlsx_files[-1]
DATA_REF = os.path.basename(XLSX_PATH).split('_')[0][:10]
print(f'Lendo: {XLSX_PATH}')

wb = openpyxl.load_workbook(XLSX_PATH)
if 'Stock Guide' not in wb.sheetnames:
    print('ERRO: Aba "Stock Guide" nao encontrada')
    sys.exit(1)

ws = wb['Stock Guide']


def safe(v):
    if v is None or v == '' or v == '-':
        return None
    try:
      return float(v)
    except (TypeError, ValueError):
      return None


def pct(v):
    s = safe(v)
    return round(s * 100, 4) if s is not None else None


def div1e6(v):
    s = safe(v)
    return round(s / 1_000_000, 3) if s is not None else None


def calc_yield_ponta(div_mes, cotacao):
    if div_mes is None or cotacao in (None, 0):
        return None
    return round((div_mes * 12 / cotacao) * 100, 4)


fiis = []
skipped = 0

for row in ws.iter_rows(min_row=6, values_only=True):
    ticker = row[1]
    segmento = row[6]
    if not ticker or not segmento:
        continue

    pvp = safe(row[14])
    dy_ltm = safe(row[16])
    vp = safe(row[13])

    if pvp is None or dy_ltm is None or vp is None:
        skipped += 1
        continue

    nome = row[3] if row[3] and row[3] != '-' else ticker
    gestor = str(row[4]).strip() if row[4] else None
    admin = str(row[5]).strip() if row[5] else None
    cotacao = safe(row[9])
    div_mes = safe(row[19])

    fiis.append({
        'ticker': ticker,
        'nome': nome,
        'gestor': gestor,
        'admin': admin,
        'segmento': segmento,
        'partIfix': pct(row[7]),
        'volMedio3m': div1e6(row[8]),
        'cotacao': cotacao,
        'max52s': safe(row[10]),
        'pctMax52s': pct(row[11]),
        'valorMercado': round(safe(row[12]) / 1_000_000, 2) if safe(row[12]) else None,
        'valorPatrimonial': round(vp / 1_000_000, 2),
        'pvp': round(pvp, 4),
        'pvp2025': round(safe(row[15]), 4) if safe(row[15]) is not None else None,
        'dyLTM': round(dy_ltm * 100, 2),
        'dyPonta': calc_yield_ponta(div_mes, cotacao),
        'dyAnualizado': pct(row[17]),
        'ultimaDistPct': pct(row[18]),
        'divMes': div_mes,
        'retMes': pct(row[20]),
        'retAno': pct(row[21]),
        'retLTM': pct(row[22]),
    })

fiis.sort(key=lambda f: f['valorPatrimonial'] or 0, reverse=True)

os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

header = f"""// ============================================================
// Stock Guide FII | Referencia: {DATA_REF}
// Gerado automaticamente por scripts/gerar_dados.py
// Total: {len(fiis)} FIIs com dados completos ({skipped} sem dados essenciais)
// ============================================================

const FIIS_DATA = """

js_content = header + json.dumps(fiis, ensure_ascii=False, indent=2) + ';\n'

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f'OK: {len(fiis)} FIIs gerados em {OUTPUT}')
print(f'   {skipped} FIIs ignorados (sem PVP, DY ou VP)')
